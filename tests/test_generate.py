import openpyxl
import pytest

from generate import ExportDataDictionary, MyWorksheet


class TestMyWorksheet:
    def test_excel_string_width_basic(self):
        ws = MyWorksheet()
        assert ws.excel_string_width("hello") == pytest.approx(5.5)

    def test_excel_string_width_empty(self):
        ws = MyWorksheet()
        assert ws.excel_string_width("") == pytest.approx(0.0)

    def test_excel_string_width_single_char(self):
        ws = MyWorksheet()
        assert ws.excel_string_width("X") == pytest.approx(1.1)

    def test_max_column_widths_tracking(self, tmp_path):
        path = str(tmp_path / "widths.xlsx")
        wb = ExportDataDictionary(path)
        ws = wb.add_worksheet("test")
        ws.write_string(0, 0, "short")
        ws.write_string(1, 0, "much longer")
        ws.write_string(0, 1, "col1")
        wb.close()
        assert ws.max_column_widths[0] == pytest.approx(len("much longer") * 1.1)
        assert ws.max_column_widths[1] == pytest.approx(len("col1") * 1.1)

    def test_shorter_string_does_not_replace_max(self, tmp_path):
        path = str(tmp_path / "no_shrink.xlsx")
        wb = ExportDataDictionary(path)
        ws = wb.add_worksheet("test")
        ws.write_string(0, 0, "a long string")
        ws.write_string(1, 0, "hi")
        wb.close()
        assert ws.max_column_widths[0] == pytest.approx(len("a long string") * 1.1)


class TestExportDataDictionary:
    def test_add_worksheet_returns_myworksheet(self, tmp_path):
        path = str(tmp_path / "ws.xlsx")
        wb = ExportDataDictionary(path)
        ws = wb.add_worksheet("Sheet1")
        assert isinstance(ws, MyWorksheet)
        wb.close()

    def test_close_produces_valid_xlsx(self, tmp_path):
        path = str(tmp_path / "close.xlsx")
        wb = ExportDataDictionary(path)
        ws = wb.add_worksheet("test")
        ws.write_string(0, 0, "some text")
        ws.write_string(0, 1, "longer text here")
        wb.close()

        workbook = openpyxl.load_workbook(path)
        assert "test" in workbook.sheetnames
        workbook.close()

    def test_write_header(self, tmp_path):
        path = str(tmp_path / "header.xlsx")
        wb = ExportDataDictionary(path)
        wb._worksheet = wb.add_worksheet("test")
        wb._row = 0
        wb._col = 0
        wb.write_header(["COL_A", "COL_B", "COL_C"])
        wb.close()

        workbook = openpyxl.load_workbook(path)
        ws = workbook.active
        assert ws.cell(1, 1).value == "COL_A"
        assert ws.cell(1, 2).value == "COL_B"
        assert ws.cell(1, 3).value == "COL_C"
        assert ws.cell(1, 1).font.bold is True
        workbook.close()

    def test_create_table_writes_table_name_row(self, tmp_path, single_table_schema):
        path = str(tmp_path / "table.xlsx")
        wb = ExportDataDictionary(path)
        wb._worksheet = wb.add_worksheet("test")
        wb._row = 0
        wb._col = 0
        wb.create_table("users", single_table_schema["users"])
        wb.close()

        workbook = openpyxl.load_workbook(path)
        ws = workbook.active
        assert ws.cell(1, 1).value == "Table:"
        assert ws.cell(1, 2).value == "users"
        workbook.close()

    def test_create_table_writes_description_row(self, tmp_path, single_table_schema):
        path = str(tmp_path / "desc.xlsx")
        wb = ExportDataDictionary(path)
        wb._worksheet = wb.add_worksheet("test")
        wb._row = 0
        wb._col = 0
        wb.create_table("users", single_table_schema["users"])
        wb.close()

        workbook = openpyxl.load_workbook(path)
        ws = workbook.active
        assert ws.cell(2, 1).value == "Description:"
        workbook.close()

    def test_create_table_writes_header_row(self, tmp_path, single_table_schema):
        path = str(tmp_path / "hdr.xlsx")
        wb = ExportDataDictionary(path)
        wb._worksheet = wb.add_worksheet("test")
        wb._row = 0
        wb._col = 0
        wb.create_table("users", single_table_schema["users"])
        wb.close()

        workbook = openpyxl.load_workbook(path)
        ws = workbook.active
        headers = [ws.cell(3, c).value for c in range(1, 8)]
        assert headers == [
            "COLUMN_ID",
            "COLUMN_NAME",
            "DESCRIPTION",
            "DATA_TYPE",
            "DATA_LENGTH",
            "NULLABLE",
            "Key Type",
        ]
        workbook.close()

    def test_create_table_writes_data_rows(self, tmp_path, single_table_schema):
        path = str(tmp_path / "data.xlsx")
        wb = ExportDataDictionary(path)
        wb._worksheet = wb.add_worksheet("test")
        wb._row = 0
        wb._col = 0
        wb.create_table("users", single_table_schema["users"])
        wb.close()

        workbook = openpyxl.load_workbook(path)
        ws = workbook.active
        # Row 4 = first data row (1-indexed in openpyxl)
        assert ws.cell(4, 1).value == "1"
        assert ws.cell(4, 2).value == "id"
        assert ws.cell(4, 3).value == "Primary key"
        assert ws.cell(4, 4).value == "int"
        assert ws.cell(4, 5).value == ""
        assert ws.cell(4, 6).value == "NO"
        assert ws.cell(4, 7).value == "PRI"
        # Row 5 = second data row
        assert ws.cell(5, 1).value == "2"
        assert ws.cell(5, 2).value == "email"
        workbook.close()

    def test_create_table_advances_row_with_gap(self, tmp_path, single_table_schema):
        path = str(tmp_path / "gap.xlsx")
        wb = ExportDataDictionary(path)
        wb._worksheet = wb.add_worksheet("test")
        wb._row = 0
        wb._col = 0
        wb.create_table("users", single_table_schema["users"])
        # table_name row (1) + description row (1) + header row (1) +
        # data rows (2) + gap (2) = 7
        assert wb._row == 7
        wb.close()

    def test_generate_xlsx_simple_single_sheet(self, tmp_path, single_table_schema):
        path = str(tmp_path / "simple.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx_simple(single_table_schema)

        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames == ["Data Dictionary"]
        workbook.close()

    def test_generate_xlsx_simple_all_tables_on_one_sheet(self, tmp_path, multi_prefix_schema):
        path = str(tmp_path / "simple_multi.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx_simple(multi_prefix_schema)

        workbook = openpyxl.load_workbook(path)
        ws = workbook["Data Dictionary"]
        table_names = []
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] == "Table:":
                table_names.append(row[1])
        assert table_names == ["app_users", "app_roles", "blog_posts", "blog_comments"]
        workbook.close()

    def test_generate_xlsx_creates_sheets_per_prefix(self, tmp_path, multi_prefix_schema):
        path = str(tmp_path / "multi.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx(multi_prefix_schema)

        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames == ["app", "blog"]
        workbook.close()

    def test_generate_xlsx_tables_on_correct_sheets(self, tmp_path, multi_prefix_schema):
        path = str(tmp_path / "correct_sheets.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx(multi_prefix_schema)

        workbook = openpyxl.load_workbook(path)
        app_tables = []
        for row in workbook["app"].iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] == "Table:":
                app_tables.append(row[1])
        assert app_tables == ["app_users", "app_roles"]

        blog_tables = []
        for row in workbook["blog"].iter_rows(min_col=1, max_col=2, values_only=True):
            if row[0] == "Table:":
                blog_tables.append(row[1])
        assert blog_tables == ["blog_posts", "blog_comments"]
        workbook.close()

    def test_generate_xlsx_single_prefix_single_sheet(self, tmp_path):
        col = {
            "ordinal": 1,
            "column_name": "id",
            "column_type": "int",
            "max_length": "",
            "is_nullable": "NO",
            "extra": "",
            "column_comment": "",
            "column_default": "",
        }
        data = {"api_users": [col], "api_tokens": [col]}
        path = str(tmp_path / "single_prefix.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx(data)

        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames == ["api"]
        workbook.close()

    def test_generate_xlsx_no_underscore_table_name(self, tmp_path):
        col = {
            "ordinal": 1,
            "column_name": "id",
            "column_type": "int",
            "max_length": "",
            "is_nullable": "NO",
            "extra": "",
            "column_comment": "",
            "column_default": "",
        }
        data = {"users": [col]}
        path = str(tmp_path / "no_underscore.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx(data)

        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames == ["users"]
        workbook.close()

    def test_generate_xlsx_empty_data(self, tmp_path):
        path = str(tmp_path / "empty.xlsx")
        wb = ExportDataDictionary(path)
        wb.generate_xlsx({})

        workbook = openpyxl.load_workbook(path)
        workbook.close()
